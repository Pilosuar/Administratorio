from django.http import HttpResponse
from .models import Alumno, Curso, Grupo, Inscripcion, EstatusInscripcion
from django.shortcuts import get_object_or_404
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image, PageBreak)
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
import os

#Genera un archivo Excel con los datos completos de todos los alumnos
def exportar_alumnos_excel(request):

    wb = Workbook()
    ws = wb.active
    ws.title = "Alumnos"

   #estilos
    borde = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    centrado = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True
    )

    titulo_fill = PatternFill("solid", fgColor="2F5496")

    azul_fuerte = PatternFill("solid", fgColor="2F5496")
    amarillo_fuerte = PatternFill("solid", fgColor="FFE598")
    verde_fuerte = PatternFill("solid", fgColor="A8D08D")

    azul_suave = PatternFill("solid", fgColor="9CC2E5")
    amarillo_suave = PatternFill("solid", fgColor="FFF2CC")
    verde_suave = PatternFill("solid", fgColor="E2F0D9")

    ws.merge_cells("A1:AM2")

    ws["A1"] = "REPORTE GENERAL DE ALUMNOS"

    ws["A1"].fill = titulo_fill
    ws["A1"].font = Font(
        bold=True,
        color="FFFFFF",
        size=16
    )

    ws["A1"].alignment = centrado

    # datos alumno
    ws.merge_cells("A4:I4")
    ws["A4"] = "DATOS DEL ALUMNO"
    ws["A4"].fill = azul_fuerte
    ws["A4"].font = Font(color="FFFFFF", bold=True)
    ws["A4"].alignment = centrado

    # datos médicos
    ws.merge_cells("J4:N4")
    ws["J4"] = "DATOS MÉDICOS"
    ws["J4"].fill = amarillo_fuerte
    ws["J4"].font = Font(bold=True)
    ws["J4"].alignment = centrado

    # datos madre
    ws.merge_cells("O4:T4")
    ws["O4"] = "DATOS DE LA MADRE"
    ws["O4"].fill = verde_fuerte
    ws["O4"].font = Font(bold=True)
    ws["O4"].alignment = centrado

    # datos padre
    ws.merge_cells("U4:Z4")
    ws["U4"] = "DATOS DEL PADRE"
    ws["U4"].fill = azul_fuerte
    ws["U4"].font = Font(color="FFFFFF", bold=True)
    ws["U4"].alignment = centrado

    # datos escolares
    ws.merge_cells("AA4:AE4")
    ws["AA4"] = "DATOS ESCOLARES"
    ws["AA4"].fill = verde_fuerte
    ws["AA4"].font = Font(bold=True)
    ws["AA4"].alignment = centrado

    # datos de referencia
    ws.merge_cells("AF4:AI4")
    ws["AF4"] = "DATOS DE REFERENCIA"
    ws["AF4"].fill = amarillo_fuerte
    ws["AF4"].font = Font(bold=True)
    ws["AF4"].alignment = centrado

    # control interno
    ws.merge_cells("AJ4:AM4")
    ws["AJ4"] = "DATOS DE CONTROL INTERNO"
    ws["AJ4"].fill = azul_fuerte
    ws["AJ4"].font = Font(color="FFFFFF", bold=True)
    ws["AJ4"].alignment = centrado

    encabezados = [

        # datos alumno
        "ID",
        "Nombre",
        "Apellido Paterno",
        "Apellido Materno",
        "Fecha Nacimiento",
        "Edad",
        "Teléfono",
        "Tipo Contacto",
        "Correo",

        #datos médicos
        "Alergias",
        "Detalle Alergias",
        "Alergia Medicamento",
        "Detalle Medicamento",
        "Tipo Sangre",

        #datos madre
        "Nombre",
        "Apellido Paterno",
        "Apellido Materno",
        "Teléfono",
        "Tipo Contacto",
        "Correo",

        #datos padre
        "Nombre",
        "Apellido Paterno",
        "Apellido Materno",
        "Teléfono",
        "Tipo Contacto",
        "Correo",

        #datos escolares
        "Escuela",
        "Nivel Escolar",
        "Año Escolar",
        "Promedio",
        "Cursos",

        #datos referencia
        "Persona Recoge",
        "Teléfono",
        "Tipo Contacto",
        "Correo",

        #control interno
        "Estatus",
        "Observaciones",
        "Fecha Creación",
        "Fecha Actualización",
    ]

    fila_encabezados = 5

    for col, texto in enumerate(encabezados, start=1):

        celda = ws.cell(fila_encabezados, col)
        celda.value = texto
        celda.border = borde
        celda.alignment = centrado
        celda.font = Font(bold=True)

        if col <= 9:
            celda.fill = azul_suave

        elif col <= 14:
            celda.fill = amarillo_suave

        elif col <= 20:
            celda.fill = verde_suave

        elif col <= 26:
            celda.fill = azul_suave

        elif col <= 31:
            celda.fill = verde_suave

        elif col <= 35:
            celda.fill = amarillo_suave

        else:
            celda.fill = azul_suave

    fila = 6

    alumnos = Alumno.objects.prefetch_related(
        "inscripciones__grupo__curso"
    ).order_by("id_alumno")

    for alumno in alumnos:

        cursos = ", ".join(
            ins.grupo.curso.nombre
            for ins in alumno.inscripciones.all()
        )

        datos = [

            alumno.id_alumno,
            alumno.nombre,
            alumno.apellido_paterno,
            alumno.apellido_materno,
            alumno.fecha_nacimiento,
            alumno.edad_calculada,
            alumno.contacto_alumno,
            alumno.contacto_alumno_tipo,
            alumno.correo_electronico,

            "Sí" if alumno.alergias else "No",
            alumno.alergias_detalle,
            "Sí" if alumno.alergico_medicamento else "No",
            alumno.alergico_medicamento_detalle,
            alumno.tipo_sangre,

            alumno.madre_nombre,
            alumno.madre_apellido_paterno,
            alumno.madre_apellido_materno,
            alumno.madre_telefono,
            alumno.madre_contacto_tipo,
            alumno.madre_correo,

            alumno.padre_nombre,
            alumno.padre_apellido_paterno,
            alumno.padre_apellido_materno,
            alumno.padre_telefono,
            alumno.padre_contacto_tipo,
            alumno.padre_correo,

            alumno.escuela,
            alumno.grado,
            alumno.año,
            alumno.promedio_anterior,
            cursos,

            alumno.persona_recoge,
            alumno.contacto_otro,
            alumno.contacto_otro_tipo,
            alumno.otro_correo,

            alumno.estatus,
            alumno.observaciones,
            alumno.fecha_creacion.strftime("%d/%m/%Y"),
            alumno.fecha_actualizacion.strftime("%d/%m/%Y"),
        ]

        for col, valor in enumerate(datos, start=1):

            celda = ws.cell(fila, col)
            celda.value = valor
            celda.border = borde
            celda.alignment = Alignment(
                vertical="center",
                wrap_text=True
            )

        fila += 1

    for columna in ws.columns:

        longitud = 0
        letra = get_column_letter(columna[0].column)

        for celda in columna:

            try:
                if celda.value:
                    longitud = max(
                        longitud,
                        len(str(celda.value))
                    )
            except:
                pass

        ws.column_dimensions[letra].width = min(
            longitud + 5,
            40
        )

    ws.freeze_panes = "A6"

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response["Content-Disposition"] = (
        'attachment; filename="Reporte_General_Alumnos.xlsx"'
    )

    wb.save(response)

    return response

#Genera un PDF con los datos completos del alumno
def descargar_pdf_alumno(request, id_alumno):

    alumno = get_object_or_404(
        Alumno,
        pk=id_alumno
    )

    response = HttpResponse(
        content_type='application/pdf'
    )

    response[
        'Content-Disposition'
    ] = f'attachment; filename="Alumno_{alumno.id_alumno}.pdf"'

    doc = SimpleDocTemplate(
        response,
        rightMargin=30,
        leftMargin=30,
        topMargin=30,
        bottomMargin=30
    )

    elementos = []
    estilos = getSampleStyleSheet()


    titulo = Paragraph(
        f"<b>EXPEDIENTE DEL ALUMNO</b>",
        estilos["Title"]
    )

    elementos.append(titulo)
    elementos.append(Spacer(1, 10))


    if alumno.foto:

        try:

            imagen = Image(
                alumno.foto.path,
                width=4*cm,
                height=4*cm
            )

            elementos.append(imagen)
            elementos.append(Spacer(1, 15))

        except:
            pass

    def agregar_titulo(texto, color):

        tabla = Table(
            [[texto]],
            colWidths=[16*cm]
        )

        tabla.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,-1), color),
            ('TEXTCOLOR', (0,0), (-1,-1), colors.white),
            ('FONTNAME', (0,0), (-1,-1), 'Helvetica-Bold'),
            ('ALIGN', (0,0), (-1,-1), 'CENTER')
        ]))

        elementos.append(tabla)
        elementos.append(Spacer(1, 5))

    def agregar_tabla(datos):

        tabla = Table(
            datos,
            colWidths=[5*cm, 11*cm]
        )

        tabla.setStyle(TableStyle([
            ('GRID',(0,0),(-1,-1),1,colors.black),
            ('BACKGROUND',(0,0),(0,-1),colors.lightgrey),
            ('FONTNAME',(0,0),(0,-1),'Helvetica-Bold'),
            ('VALIGN',(0,0),(-1,-1),'MIDDLE')
        ]))

        elementos.append(tabla)
        elementos.append(Spacer(1, 10))

    agregar_titulo(
        "DATOS PERSONALES",
        colors.HexColor("#2F5496")
    )

    agregar_tabla([
        ["ID", alumno.id_alumno],
        ["Nombre", alumno.nombre],
        ["Apellido paterno", alumno.apellido_paterno],
        ["Apellido materno", alumno.apellido_materno],
        ["Edad", alumno.edad_calculada],
        ["Fecha nacimiento", alumno.fecha_nacimiento.strftime("%d/%m/%Y")],
        ["Teléfono", alumno.contacto_alumno],
        ["Correo", alumno.correo_electronico],
        ["Estatus", alumno.estatus],
    ])

    agregar_titulo(
        "DATOS MÉDICOS",
        colors.HexColor("#9CC2E5")
    )

    agregar_tabla([
        ["Tipo sangre", alumno.tipo_sangre],
        ["Alergias", "Sí" if alumno.alergias else "No"],
        ["Detalle", alumno.alergias_detalle],
        ["Alergia medicamento", "Sí" if alumno.alergico_medicamento else "No"],
        ["Medicamento", alumno.alergico_medicamento_detalle],
    ])

    agregar_titulo(
        "DATOS DE LA MADRE",
        colors.HexColor("#FFE699")
    )

    agregar_tabla([
        ["Nombre", alumno.madre_nombre],
        ["Apellido paterno", alumno.madre_apellido_paterno],
        ["Apellido materno", alumno.madre_apellido_materno],
        ["Teléfono", alumno.madre_telefono],
        ["Correo", alumno.madre_correo],
    ])

    agregar_titulo(
        "DATOS DEL PADRE",
        colors.HexColor("#A9D18E")
    )

    agregar_tabla([
        ["Nombre", alumno.padre_nombre],
        ["Apellido paterno", alumno.padre_apellido_paterno],
        ["Apellido materno", alumno.padre_apellido_materno],
        ["Teléfono", alumno.padre_telefono],
        ["Correo", alumno.padre_correo],
    ])

    agregar_titulo(
        "DATOS ESCOLARES",
        colors.HexColor("#2F5496")
    )

    agregar_tabla([
        ["Escuela", alumno.escuela],
        ["Nivel escolar", alumno.get_grado_display()],
        ["Año escolar", alumno.get_año_display()],
        ["Promedio", alumno.promedio_anterior],
    ])

    agregar_titulo(
        "DATOS DE REFERENCIA",
        colors.HexColor("#FFE699")
    )

    agregar_tabla([
        ["Persona que recoge", alumno.persona_recoge],
        ["Teléfono", alumno.contacto_otro],
        ["Correo", alumno.otro_correo],
    ])

    agregar_titulo(
        "DATOS INTERNOS",
        colors.HexColor("#A9D18E")
    )

    agregar_tabla([
        ["Fecha creación", alumno.fecha_creacion.strftime("%d/%m/%Y %H:%M")],
        ["Última actualización", alumno.fecha_actualizacion.strftime("%d/%m/%Y %H:%M")],
        ["Observaciones", alumno.observaciones],
    ])

    agregar_titulo(
        "INSCRIPCIONES",
        colors.HexColor("#9CC2E5")
    )

    datos_inscripciones = [
        [
            "Curso",
            "Grupo",
            "Fecha",
            "Estatus"
        ]
    ]

    for ins in alumno.inscripciones.all():

        datos_inscripciones.append([
            ins.grupo.curso.nombre,
            ins.grupo.nombre,
            ins.fecha_inscripcion.strftime("%d/%m/%Y"),
            ins.estatus
        ])

    tabla_ins = Table(datos_inscripciones)

    tabla_ins.setStyle(TableStyle([
        ('GRID',(0,0),(-1,-1),1,colors.black),
        ('BACKGROUND',(0,0),(-1,0),colors.HexColor("#2F5496")),
        ('TEXTCOLOR',(0,0),(-1,0),colors.white),
        ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold')
    ]))

    elementos.append(tabla_ins)

    doc.build(elementos)

    return response


def exportar_inscripciones_excel(request):

    wb = Workbook()
    ws = wb.active
    ws.title = "Inscripciones"

    borde = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    centrado = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True
    )

    titulo_fill = PatternFill(
        "solid",
        fgColor="2F5496"
    )

    encabezado_fill = PatternFill(
        "solid",
        fgColor="9CC2E5"
    )

    ws.merge_cells("A1:K2")

    ws["A1"] = "REPORTE GENERAL DE INSCRIPCIONES"

    ws["A1"].fill = titulo_fill
    ws["A1"].font = Font(
        bold=True,
        color="FFFFFF",
        size=16
    )

    ws["A1"].alignment = centrado

    encabezados = [
        "ID",
        "Alumno",
        "Curso",
        "Grupo",
        "Horario",
        "Fecha Inscripción",
        "Contrato PDF",
        "Estatus",
        "Fecha Baja",
        "Fecha Creación",
        "Fecha Actualización",
        "Observaciones"
    ]

    fila_encabezados = 4

    for col, texto in enumerate(encabezados, start=1):

        celda = ws.cell(fila_encabezados, col)

        celda.value = texto
        celda.fill = encabezado_fill
        celda.font = Font(bold=True)
        celda.border = borde
        celda.alignment = centrado

    fila = 5

    inscripciones = (
        Inscripcion.objects
        .select_related(
            "alumno",
            "grupo",
            "grupo__curso"
        )
        .order_by("id")
    )

    for inscripcion in inscripciones:

        datos = [
            inscripcion.id,
            str(inscripcion.alumno),
            inscripcion.grupo.curso.nombre,
            inscripcion.grupo.nombre,
            inscripcion.grupo.horario,
            inscripcion.fecha_inscripcion.strftime("%d/%m/%Y"),
            "Sí" if inscripcion.contrato_pdf else "No",
            inscripcion.estatus,
            (
                inscripcion.fecha_baja.strftime("%d/%m/%Y")
                if inscripcion.fecha_baja
                else ""
            ),
            inscripcion.fecha_creacion.strftime("%d/%m/%Y"),
            inscripcion.fecha_actualizacion.strftime("%d/%m/%Y"),
            inscripcion.observaciones,
        ]

        for col, valor in enumerate(datos, start=1):

            celda = ws.cell(fila, col)

            celda.value = valor
            celda.border = borde
            celda.alignment = Alignment(
                vertical="center",
                wrap_text=True
            )

        fila += 1

    for columna in ws.columns:

        longitud = 0

        try:
            letra = get_column_letter(
                columna[0].column
            )
        except:
            continue

        for celda in columna:

            try:
                if celda.value:
                    longitud = max(
                        longitud,
                        len(str(celda.value))
                    )
            except:
                pass

        ws.column_dimensions[letra].width = min(
            longitud + 5,
            50
        )

    ws.freeze_panes = "A5"

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response["Content-Disposition"] = (
        'attachment; filename="Reporte_Inscripciones.xlsx"'
    )

    wb.save(response)

    return response

#Genera un archivo excel de los cursos registrados
def exportar_cursos_excel(request):

    wb = Workbook()
    ws = wb.active
    ws.title = "Cursos"

    borde = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    centrado = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True
    )

    titulo_fill = PatternFill(
        "solid",
        fgColor="2F5496"
    )

    encabezado_fill = PatternFill(
        "solid",
        fgColor="9CC2E5"
    )

    ws.merge_cells("A1:C2")

    ws["A1"] = "REPORTE GENERAL DE CURSOS"

    ws["A1"].fill = titulo_fill
    ws["A1"].font = Font(
        bold=True,
        color="FFFFFF",
        size=16
    )

    ws["A1"].alignment = centrado

    encabezados = [
        "ID",
        "Nombre del Curso",
        "Descripción"
    ]

    fila_encabezados = 4

    for col, texto in enumerate(encabezados, start=1):

        celda = ws.cell(fila_encabezados, col)

        celda.value = texto
        celda.fill = encabezado_fill
        celda.font = Font(bold=True)
        celda.border = borde
        celda.alignment = centrado

    fila = 5

    cursos = Curso.objects.all().order_by("id")

    for curso in cursos:

        datos = [
            curso.id,
            curso.nombre,
            curso.descripcion
        ]

        for col, valor in enumerate(datos, start=1):

            celda = ws.cell(fila, col)

            celda.value = valor
            celda.border = borde
            celda.alignment = Alignment(
                vertical="center",
                wrap_text=True
            )

        fila += 1

    for columna in ws.columns:

        longitud = 0

        try:
            letra = get_column_letter(
                columna[0].column
            )
        except:
            continue

        for celda in columna:

            try:
                if celda.value:
                    longitud = max(
                        longitud,
                        len(str(celda.value))
                    )
            except:
                pass

        ws.column_dimensions[letra].width = min(
            longitud + 5,
            50
        )

    ws.freeze_panes = "A5"

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response["Content-Disposition"] = (
        'attachment; filename="Reporte_Cursos.xlsx"'
    )

    wb.save(response)

    return response


#Genera un archivo excel del curso y de los alumnos que estan o estuvieron inscritos
def exportar_curso_excel(request, id):

    curso = get_object_or_404(
        Curso,
        pk=id
    )

    inscripciones = (
        Inscripcion.objects
        .select_related(
            "alumno",
            "grupo"
        )
        .filter(
            grupo__curso=curso
        )
        .order_by(
            "fecha_inscripcion"
        )
    )

    wb = Workbook()
    ws = wb.active

    nombre_hoja = curso.nombre[:31]
    ws.title = nombre_hoja

    borde = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    centrado = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True
    )

    titulo_fill = PatternFill(
        "solid",
        fgColor="2F5496"
    )

    encabezado_fill = PatternFill(
        "solid",
        fgColor="9CC2E5"
    )

    ws.merge_cells("A1:F2")

    ws["A1"] = (
        f"REPORTE DE ALUMNOS DEL CURSO: "
        f"{curso.nombre.upper()}"
    )

    ws["A1"].fill = titulo_fill

    ws["A1"].font = Font(
        bold=True,
        color="FFFFFF",
        size=16
    )

    ws["A1"].alignment = centrado


    encabezados = [
        "ID Alumno",
        "Nombre Completo",
        "Grupo",
        "Fecha Inicio",
        "Fecha Fin",
        "Estatus"
    ]

    fila_encabezados = 4

    for col, texto in enumerate(encabezados, start=1):

        celda = ws.cell(
            fila_encabezados,
            col
        )

        celda.value = texto
        celda.fill = encabezado_fill
        celda.font = Font(bold=True)
        celda.border = borde
        celda.alignment = centrado

    fila = 5

    for inscripcion in inscripciones:

        alumno = inscripcion.alumno

        nombre_completo = (
            f"{alumno.nombre} "
            f"{alumno.apellido_paterno} "
            f"{alumno.apellido_materno}"
        )

        datos = [

            alumno.id_alumno,

            nombre_completo,

            inscripcion.grupo.nombre,

            inscripcion.fecha_inscripcion.strftime(
                "%d/%m/%Y"
            ),

            (
                inscripcion.fecha_baja.strftime(
                    "%d/%m/%Y"
                )
                if inscripcion.fecha_baja
                else ""
            ),

            inscripcion.estatus,
        ]

        for col, valor in enumerate(
            datos,
            start=1
        ):

            celda = ws.cell(
                fila,
                col
            )

            celda.value = valor
            celda.border = borde
            celda.alignment = Alignment(
                vertical="center",
                wrap_text=True
            )

        fila += 1

    for columna in ws.columns:

        try:
            letra = get_column_letter(
                columna[0].column
            )
        except:
            continue

        longitud = 0

        for celda in columna:

            try:
                if celda.value:
                    longitud = max(
                        longitud,
                        len(str(celda.value))
                    )
            except:
                pass

        ws.column_dimensions[
            letra
        ].width = min(
            longitud + 5,
            50
        )

    ws.freeze_panes = "A5"

    nombre_archivo = (
        f"Curso_{curso.nombre}.xlsx"
    )

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response["Content-Disposition"] = (
        f'attachment; filename="{nombre_archivo}"'
    )

    wb.save(response)

    return response

def exportar_grupos_excel(request):

    wb = Workbook()
    ws = wb.active
    ws.title = "Grupos"

    borde = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    centrado = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True
    )

    titulo_fill = PatternFill(
        "solid",
        fgColor="2F5496"
    )

    encabezado_fill = PatternFill(
        "solid",
        fgColor="9CC2E5"
    )

    ws.merge_cells("A1:E2")

    ws["A1"] = "REPORTE GENERAL DE GRUPOS"

    ws["A1"].fill = titulo_fill

    ws["A1"].font = Font(
        bold=True,
        color="FFFFFF",
        size=16
    )

    ws["A1"].alignment = centrado

    encabezados = [
        "ID",
        "Grupo",
        "Curso",
        "Horario",
        "Cupo"
    ]

    fila_encabezados = 4

    for col, texto in enumerate(encabezados, start=1):

        celda = ws.cell(fila_encabezados, col)

        celda.value = texto
        celda.fill = encabezado_fill
        celda.font = Font(bold=True)
        celda.border = borde
        celda.alignment = centrado

    fila = 5

    grupos = (
        Grupo.objects
        .select_related("curso")
        .order_by("id")
    )

    for grupo in grupos:

        datos = [
            grupo.id,
            grupo.nombre,
            grupo.curso.nombre,
            grupo.horario,
            grupo.cupo
        ]

        for col, valor in enumerate(datos, start=1):

            celda = ws.cell(fila, col)

            celda.value = valor
            celda.border = borde
            celda.alignment = Alignment(
                vertical="center",
                wrap_text=True
            )

        fila += 1

    for columna in ws.columns:

        try:
            letra = get_column_letter(
                columna[0].column
            )
        except:
            continue

        longitud = 0

        for celda in columna:

            try:
                if celda.value:
                    longitud = max(
                        longitud,
                        len(str(celda.value))
                    )
            except:
                pass

        ws.column_dimensions[
            letra
        ].width = min(
            longitud + 5,
            50
        )

    ws.freeze_panes = "A5"

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response["Content-Disposition"] = (
        'attachment; filename="Reporte_Grupos.xlsx"'
    )

    wb.save(response)

    return response